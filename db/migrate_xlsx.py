#!/usr/bin/env python3
"""
Migrate Exotiq_Operator_CRM_Master.xlsx into the Exotiq SQLite database.

Usage:
    python db/migrate_xlsx.py [path/to/Exotiq_Operator_CRM_Master.xlsx]

If no path is given, looks for the file at the project root.
"""

import json
import re
import sys
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
import openpyxl

ROOT = Path(__file__).parent.parent
DEFAULT_XLSX_PATH = ROOT / "Exotiq_Operator_CRM_Master.xlsx"
DB_PATH = ROOT / "db/exotiq.db"

# List of tabs that contain leads
LEAD_MARKET_TABS = [
    'Miami Operators', 'Phoenix Scottsdale', 'Dallas Fort Worth', 'Chicago', 
    'Atlanta', 'NYC', 'Las Vegas', 'Los Angeles', 'SF Bay Area', 'DC DMV'
]

MARKET_ABBREVS = {
    'Miami Operators': 'mia', 'Phoenix Scottsdale': 'phx', 'Dallas Fort Worth': 'dfw',
    'Chicago': 'chi', 'Atlanta': 'atl', 'NYC': 'nyc', 'Las Vegas': 'lvs',
    'Los Angeles': 'lax', 'SF Bay Area': 'sfb', 'DC DMV': 'dcv'
}

MARKET_NAMES = {
    'Miami Operators': 'Miami', 'Phoenix Scottsdale': 'Phoenix/Scottsdale', 'Dallas Fort Worth': 'Dallas/Fort Worth',
    'Chicago': 'Chicago', 'Atlanta': 'Atlanta', 'NYC': 'NYC', 'Las Vegas': 'Las Vegas',
    'Los Angeles': 'Los Angeles', 'SF Bay Area': 'SF Bay Area', 'DC DMV': 'DC/DMV'
}

MARKET_NORMALIZE = {
    'tampa': 'Miami', 'hollywood': 'Miami', 'fort lauderdale': 'Miami', 
    'boca raton': 'Miami', 'miami beach': 'Miami', 'miami': 'Miami',
    'tampa, la, oc': 'Multi-Market',
    'nyc, nj, philly & dc': 'NYC',
}

# --- Utility functions ---

def is_phone(val):
    if not val: return False
    return len(re.sub(r'[^0-9+]', '', str(val))) >= 10

def clean_str(val):
    if val is None: return None
    s = str(val).strip()
    return s if s and s.lower() != 'none' else None

def parse_fleet_size(val):
    if not val: return None
    nums = re.findall(r'(\d+)', str(val))
    return int(nums[0]) if nums else None

# --- Main migration logic ---

def run_migration(xlsx_path):
    print(f'Opening workbook: {xlsx_path.name}')
    
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    now = datetime.now(timezone.utc).isoformat()
    
    # Wipe and reinit DB
    conn.execute('DELETE FROM leads')
    conn.execute('DELETE FROM dm_drafts')
    conn.execute('DELETE FROM activity_log WHERE type != "system"')
    conn.commit()
    print('DB tables cleared for clean import.')

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    
    total, warnings, market_counts = 0, [], {}

    for tab_name in wb.sheetnames:
        if tab_name not in LEAD_MARKET_TABS:
            print(f'  Skipping non-lead tab: {tab_name}')
            continue
        
        print(f'  Importing leads from: {tab_name}')
        ws = wb[tab_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2: continue
        
        headers = [str(h).strip() if h else '' for h in rows[0]]
        default_market = MARKET_NAMES.get(tab_name, tab_name)
        abbrev = MARKET_ABBREVS.get(tab_name, 'unk')
        seq = 0
        
        for row in rows[1:]:
            if not any(c for c in row if c is not None and str(c).strip()):
                continue
            
            def get(col_name):
                try: 
                    idx = headers.index(col_name)
                    return clean_str(row[idx])
                except (ValueError, IndexError): return None
            
            company = get('Company')
            if not company: continue
            
            if company.startswith('http'):
                warnings.append(f'[{tab_name}] Company is URL: {company}')
                company = re.sub(r'https?://(www\.)?', '', company).split('/')[0].split('.')[0].title()
            
            seq += 1
            lead_id = f'lead_{abbrev}_{seq:03d}'
            
            raw_market = get('Market') or default_market
            market = MARKET_NORMALIZE.get(raw_market.lower().strip(), raw_market) if raw_market else default_market
            
            email = get('Email')
            phone = get('Company Phone')
            if email and is_phone(email):
                if not phone: phone = email
                email = None
                warnings.append(f'[{tab_name}] {company}: Email field remapped to phone')
            
            client_review = get('Client Review (Y/N)')
            confidence = 'CONFIRMED' if client_review and client_review.upper() == 'Y' else 'ESTIMATED'
            
            draft_dm = get('Draft DM')
            approved_dm = get('Approved/Sent DM')
            
            lead = {
                'id': lead_id, 'company': company,
                'contact_first_name': get('First Name'), 'contact_last_name': get('Last Name'),
                'contact_title': get('Title'), 'contact_email': email, 'contact_phone': phone,
                'contact_linkedin': get('LinkedIn URL (personal)'), 'contact_ig_personal': get('IG Handle (personal)'),
                'contact_first_name_source': 'manual', 'contact_last_name_source': 'manual',
                'contact_title_source': 'manual', 'contact_email_source': 'manual' if email else None,
                'contact_phone_source': 'manual' if phone else None,
                'contact_first_name_confidence': confidence, 'contact_last_name_confidence': confidence,
                'contact_title_confidence': confidence, 'contact_email_confidence': confidence if email else None,
                'contact_phone_confidence': confidence if phone else None,
                'company_ig_handle': get('IG Handle (company)'),
                'company_address': f"{get('City') or ''}, {get('State') or ''}".strip(', '),
                'fleet_size': parse_fleet_size(get('Fleet Size')),
                'fleet_size_source': 'manual',
                'fleet_size_confidence': confidence,
                'scoring_score': int(float(get('Lead Score'))) if get('Lead Score') and str(get('Lead Score')).replace('.', '', 1).isdigit() else None,
                'scoring_confidence': 'HIGH' if client_review and client_review.upper() == 'Y' else 'MEDIUM',
                'scoring_rationale': get('Enrichment Notes'), 'scoring_scored_at': now,
                'outreach_status': get('Status') or 'New',
                'outreach_dm_draft': approved_dm or draft_dm,
                'outreach_template_used': get('DM Code'),
                'outreach_client_review': client_review,
                'outreach_approval_status': 'APPROVED' if approved_dm else ('PENDING' if draft_dm else None),
                'outreach_dm1_sent': clean_str(get('DM1 Sent Date')),
                'outreach_response_received': bool(get('Response Received') and get('Response Received').upper() in ('Y', 'YES', 'TRUE')),
                'market': market, 'lead_source': get('Lead Source') or 'Manual',
                'notes': ' | '.join(filter(None, [get('Notes'), get('Client Notes'), get('Response Notes')])),
                'created_at': now, 'updated_at': now,
            }
            
            cols = [k for k in lead.keys() if lead[k] is not None]
            conn.execute(f'INSERT OR REPLACE INTO leads ({", ".join(cols)}) VALUES ({", ".join(["?"]*len(cols))})', [lead[k] for k in cols])
            
            total += 1
            market_counts[market] = market_counts.get(market, 0) + 1

    conn.commit()
    conn.execute('INSERT INTO activity_log (timestamp, type, description, source, agent) VALUES (?, ?, ?, ?, ?)',
                 (now, 'migration', f'Clean master CRM import: {total} leads across {len(market_counts)} markets', 'manual', 'saul'))
    conn.commit()
    conn.close()
    wb.close()
    
    print_report(total, market_counts, warnings)

def print_report(total, market_counts, warnings):
    print('='*40)
    print('  Exotiq Clean Migration Report')
    print('='*40)
    print(f'Total leads imported: {total}')
    print()
    for m, c in sorted(market_counts.items(), key=lambda x: -x[1]):
        print(f'  {m}: {c}')
    print()
    if warnings:
        print(f'Warnings ({len(warnings)}):')
        for w in warnings:
            print(f'  - {w}')
    print('='*40)

if __name__ == '__main__':
    path_arg = sys.argv[1] if len(sys.argv) > 1 else None
    xlsx_path = Path(path_arg) if path_arg else DEFAULT_XLSX_PATH
    if not xlsx_path.exists():
        print(f'Error: Could not find workbook at {xlsx_path}')
        sys.exit(1)
    run_migration(xlsx_path)
